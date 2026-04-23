"""
escribir_excel.py
=================
Escribe los resultados del modelo de vuelta al mismo archivo Excel.

Hojas de salida:
  - "Tramos"              → df_salida completo (fila por tramo × año)
  - "Resumen Tramos"      → df_agregado (fila por tramo)
  - "consolidado2"        → series verticales año a año
  - "Copia de Consolidado"→ series horizontales año a año (preserva fórmulas)

Uso:
    from escribir_excel import escribir_resultados, escribir_resultados_bytes
    escribir_resultados("Resultados_Modelo.xlsx", df_salida, df_agregado)
    # o, con el .xlsx en memoria (p. ej. Streamlit):
    blob = escribir_resultados_bytes(uploaded_bytes, df_salida, df_agregado)
"""

import shutil
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from modelo_gf import (
    calc_suma_simple_por_anio,
    calc_suma_producto,
    calc_min_carga_por_anio,
)


# Campos con serie desde año 1 (obras, mantenimiento, vel/carga): mismo cálculo
# para consolidado2 (vertical) y Copia de Consolidado (horizontal).
_CAMPOS_DESDE_ANIO_1: list[tuple[str, bool]] = [
    ("Obra de Renovación", True),
    ("Obra de Mejoramiento", True),
    ("Desvíos a construir", True),
    ("Conservación costo fijo anual", True),
    ("Primer Tramo de Mantenimiento", True),
    ("Segundo Tramo de Mantenimiento", True),
    ("Costo operación infraestructura del tramo", True),
    ("Velocidad", False),
    ("Carga x Eje", False),
]

# consolidado2: columna Excel (1-based) → clave en series
_COL_CONSOLIDADO2_OBRAS: list[tuple[int, str]] = [
    (4, "Obra de Renovación"),
    (5, "Obra de Mejoramiento"),
    (6, "Desvíos a construir"),
    (7, "Conservación costo fijo anual"),
    (8, "Primer Tramo de Mantenimiento"),
    (9, "Segundo Tramo de Mantenimiento"),
    (10, "Costo operación infraestructura del tramo"),
    (12, "Velocidad"),
    (13, "Carga x Eje"),
]

# Copia de Consolidado: fila Excel → nombre de campo (misma serie que arriba)
_FILA_COPIA_OBRAS: list[tuple[int, str]] = [
    (12, "Obra de Renovación"),
    (13, "Obra de Mejoramiento"),
    (14, "Desvíos a construir"),
    (17, "Conservación costo fijo anual"),
    (18, "Primer Tramo de Mantenimiento"),
    (19, "Segundo Tramo de Mantenimiento"),
    (21, "Costo operación infraestructura del tramo"),
    (27, "Velocidad"),
    (28, "Carga x Eje"),
]


def _series_consolidado_desde_df(df_salida: pd.DataFrame) -> dict:
    """
    Calcula una sola vez todas las series para consolidado2 y Copia de Consolidado.
    Claves: longitud_total, carga_prom, tnkm, vida_util, min_eje, y cada nombre
    en _CAMPOS_DESDE_ANIO_1.
    """
    longitud_total = float(df_salida[df_salida["año"] == 0]["longitud"].sum())
    out: dict = {
        "longitud_total": longitud_total,
        "carga_prom": calc_suma_producto(
            df_salida,
            "Carga proyectada por año (q) (tn) - LIMITADA",
            desde_anio_1=False,
            por_longitud=False,
        ),
        "tnkm": calc_suma_simple_por_anio(df_salida, "Carga anual TN.KM"),
        "vida_util": calc_suma_producto(
            df_salida, "% Vida útil consumida", desde_anio_1=False, por_longitud=False
        ),
        "min_eje": calc_min_carga_por_anio(df_salida, "Carga x Eje"),
    }
    for campo, por_longitud in _CAMPOS_DESDE_ANIO_1:
        out[campo] = calc_suma_producto(
            df_salida, campo, desde_anio_1=True, por_longitud=por_longitud
        )
    return out


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clear_sheet_data(ws, from_row: int, max_col: int = None):
    """Borra valores (no fórmulas) desde from_row hasta el final."""
    max_r = ws.max_row
    max_c = max_col or ws.max_column
    for row in ws.iter_rows(min_row=from_row, max_row=max_r, max_col=max_c):
        for cell in row:
            if not isinstance(cell.value, str) or not cell.value.startswith('='):
                cell.value = None


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, include_index: bool = True):
    """
    Escribe un DataFrame en la hoja desde start_row.
    Si include_index=True escribe el índice como primera columna.
    """
    if include_index:
        data = df.reset_index()
    else:
        data = df.reset_index(drop=True)

    for r_idx, row in enumerate(data.itertuples(index=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            # Convertir numpy types a Python nativos
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = None if np.isnan(val) else float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(row=r_idx, column=c_idx, value=val)


def _write_series_vertical(ws, series: pd.Series, col: int, start_row: int):
    """Escribe una Serie en una columna, empezando en start_row."""
    for r_offset, val in enumerate(series):
        if isinstance(val, (np.integer,)):
            val = int(val)
        elif isinstance(val, (np.floating,)):
            val = None if np.isnan(val) else float(val)
        elif isinstance(val, float) and np.isnan(val):
            val = None
        ws.cell(row=start_row + r_offset, column=col, value=val)


def _write_series_horizontal(ws, series: pd.Series, row: int, start_col: int):
    """Escribe una Serie en una fila, empezando en start_col."""
    for c_offset, val in enumerate(series):
        if isinstance(val, (np.integer,)):
            val = int(val)
        elif isinstance(val, (np.floating,)):
            val = None if np.isnan(val) else float(val)
        elif isinstance(val, float) and np.isnan(val):
            val = None
        ws.cell(row=row, column=start_col + c_offset, value=val)


# ---------------------------------------------------------------------------
# Hoja "Tramos" — df_salida completo
# ---------------------------------------------------------------------------

def _escribir_tramos(ws, df_salida: pd.DataFrame):
    """
    Escribe df_salida en la hoja "Tramos".
    Fila 1: encabezados. Fila 2 en adelante: datos.
    """
    # Borrar datos existentes (desde fila 2)
    _clear_sheet_data(ws, from_row=2)

    cols = df_salida.columns.tolist()

    # Encabezados en fila 1
    for c_idx, col_name in enumerate(cols, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # Datos desde fila 2
    for r_idx, row in enumerate(df_salida.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = None if np.isnan(val) else float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(row=r_idx, column=c_idx, value=val)

    print(f"  ✓ Tramos: {len(df_salida)} filas escritas")


# ---------------------------------------------------------------------------
# Hoja "Resumen Tramos" — df_agregado
# ---------------------------------------------------------------------------

def _escribir_resumen_tramos(ws, df_agregado: pd.DataFrame):
    """
    Escribe df_agregado en la hoja "Resumen Tramos".
    Fila 1: encabezados (incluyendo el índice id_tramo como "Id").
    Fila 2 en adelante: datos.
    """
    _clear_sheet_data(ws, from_row=2)

    # El índice de df_agregado es id_tramo; puede que ya exista como columna
    data = df_agregado.reset_index(drop='id_tramo' in df_agregado.columns)
    cols = data.columns.tolist()

    # Encabezados fila 1
    for c_idx, col_name in enumerate(cols, start=1):
        ws.cell(row=1, column=c_idx, value=col_name)

    # Datos desde fila 2
    for r_idx, row_data in enumerate(data.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row_data, start=1):
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = None if np.isnan(val) else float(val)
            elif isinstance(val, float) and np.isnan(val):
                val = None
            ws.cell(row=r_idx, column=c_idx, value=val)

    print(f"  ✓ Resumen Tramos: {len(df_agregado)} filas escritas")


# ---------------------------------------------------------------------------
# Hoja "consolidado2" — series verticales
# ---------------------------------------------------------------------------
#
# Layout (desde el Excel original):
#   Col A: Carga anual promedio ponderada TN  → desde año 0, ponderada
#   Col B: Carga anual TN.KM                  → simple sum, desde año 0
#   Col C: % Vida útil consumida              → desde año 0, ponderada
#   Col D: Obra de Renovación                 → desde año 1, × longitud
#   Col E: Obra de Mejoramiento               → desde año 1, × longitud
#   Col F: Desvíos a construir                → desde año 1, × longitud
#   Col G: Conservación costo fijo anual      → desde año 1, × longitud
#   Col H: Primer Tramo de Mantenimiento      → desde año 1, × longitud
#   Col I: Segundo Tramo de Mantenimiento     → desde año 1, × longitud
#   Col J: Costo operación infra tramo        → desde año 1, × longitud
#   Col K: Costo PCT                          → (reservado, 0 por ahora)
#   Col L: Velocidad máxima promedio          → desde año 1, ponderada
#   Col M: Carga x Eje ponderada              → desde año 1, ponderada
#   Col N: Carga x Eje limitante (mínima)     → desde año 1, mínima
#
# Año 0 → fila 2. Año 1 → fila 3 (columnas que empiezan en año 1).

def _escribir_consolidado2(ws, df_salida: pd.DataFrame):
    # Borrar datos (fila 2 en adelante)
    _clear_sheet_data(ws, from_row=2, max_col=20)

    # Encabezado fila 1 (ya está en el Excel, pero lo re-escribimos por si acaso)
    headers = [
        'Carga anual promedio ponderada TN',
        'Carga anual TN.KM',
        '%Vida útil consumida promedio',
        'Obra de Renovación',
        'Obra de Mejoramiento',
        'Desvíos a construir',
        'Conservación costo fijo anual',
        'Primer Tramo de Mantenimiento',
        'Segundo Tramo de Mantenimiento',
        'Costo operación infraestructura del tramo',
        'Costo PCT',
        'VELOCIDAD MÁXIMA PROMEDIO (Km/h)',
        'CARGA POR EJE PROMEDIO PONDERADA POR LONG.',
        'CARGA POR EJE LIMITANTE DE LA RED',
    ]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h)

    ser = _series_consolidado_desde_df(df_salida)
    s_carga_prom = ser["carga_prom"]
    s_tnkm = ser["tnkm"]
    s_vida_util = ser["vida_util"]

    _write_series_vertical(ws, s_carga_prom, col=1, start_row=2)  # A
    _write_series_vertical(ws, s_tnkm, col=2, start_row=2)  # B
    _write_series_vertical(ws, s_vida_util, col=3, start_row=2)  # C

    for col_idx, campo in _COL_CONSOLIDADO2_OBRAS:
        _write_series_vertical(ws, ser[campo], col=col_idx, start_row=3)

    _write_series_vertical(ws, ser["min_eje"], col=14, start_row=3)

    print(f"  ✓ consolidado2: series de {len(s_carga_prom)} años escritas")


# ---------------------------------------------------------------------------
# Hoja "Copia de Consolidado" — series horizontales
# ---------------------------------------------------------------------------
#
# Layout del Excel:
#   Fila 1, col B:  longitud total de la red
#   Fila 7:         años 0..40 en cols D..AR (D=año0, E=año1, ..., AR=año40)
#   Fila 8,  col D: Carga anual prom. ponderada TN  (año 0-40)
#   Fila 9,  col D: Carga anual TN.KM               (año 0-40)
#   Fila 10, col D: % Vida útil consumida           (año 0-40)
#   Fila 12, col E: Obra de Renovación              (año 1-40)
#   Fila 13, col E: Obra de Mejoramiento            (año 1-40)
#   Fila 14, col E: Desvíos a construir             (año 1-40)
#   Fila 17, col E: Conservación costo fijo anual   (año 1-40)
#   Fila 18, col E: Primer Tramo de Mantenimiento   (año 1-40)
#   Fila 19, col E: Segundo Tramo de Mantenimiento  (año 1-40)
#   Fila 21, col E: Costo operación infra tramo     (año 1-40)
#   Fila 27, col E: Velocidad promedio              (año 1-40)
#   Fila 28, col E: Carga x Eje ponderada           (año 1-40)
#   Fila 29, col E: Carga x Eje mínima              (año 1-40)
#
# Col D = columna 4, Col E = columna 5

_COL_D = 4   # año 0
_COL_E = 5   # año 1


def _escribir_copia_consolidado(ws, df_salida: pd.DataFrame):
    """
    Escribe la hoja «Copia de Consolidado»: series horizontales (años en columnas).
    Solo toca celdas de datos calculados; no borra fórmulas en otras columnas.
    Usa las mismas series que consolidado2 vía _series_consolidado_desde_df.
    """
    ser = _series_consolidado_desde_df(df_salida)
    longitud_total = ser["longitud_total"]
    max_anio = df_salida['año'].max()

    ws.cell(row=1, column=2, value=round(longitud_total, 2))

    for anio in range(0, max_anio + 1):
        ws.cell(row=7, column=_COL_D + anio, value=anio)

    _write_series_horizontal(ws, ser["carga_prom"], row=8, start_col=_COL_D)
    _write_series_horizontal(ws, ser["tnkm"], row=9, start_col=_COL_D)
    _write_series_horizontal(ws, ser["vida_util"], row=10, start_col=_COL_D)

    for fila, campo in _FILA_COPIA_OBRAS:
        _write_series_horizontal(ws, ser[campo], row=fila, start_col=_COL_E)

    _write_series_horizontal(ws, ser["min_eje"], row=29, start_col=_COL_E)

    print(f"  ✓ Copia de Consolidado: series escritas (longitud red = {longitud_total:.1f} km)")


def _escribir_variables_modelo(ws, variables: dict) -> None:
    """
    Actualiza la hoja "Variables del Modelo" con los valores del dict variables.
    Estructura: Col A: descripción, Col B: valor, Col C: nombre_variable
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if len(row) < 3:
            continue
        var_name_cell = row[2]  # Col C
        value_cell = row[1]     # Col B
        if var_name_cell.value and str(var_name_cell.value).strip() in variables:
            var_name = str(var_name_cell.value).strip()
            new_value = variables[var_name]
            # Convertir valor para Excel
            if isinstance(new_value, bool):
                value_cell.value = new_value
            elif isinstance(new_value, (int, float)):
                value_cell.value = new_value
            else:
                value_cell.value = str(new_value)


# ---------------------------------------------------------------------------
# Punto de entrada principal
# ---------------------------------------------------------------------------

def escribir_resultados(
    path_excel: str,
    df_salida: pd.DataFrame,
    df_agregado: pd.DataFrame,
    path_salida: str = None,
    variables: dict = None,
) -> str:
    """
    Escribe los resultados del modelo en el archivo Excel.

    Args:
        path_excel  : ruta al archivo Excel (lectura base y escritura)
        df_salida   : DataFrame con resultado por tramo × año
        df_agregado : DataFrame con resumen por tramo
        path_salida : ruta de salida (si None, sobreescribe path_excel)
        variables   : dict con variables del modelo para actualizar la hoja "Variables del Modelo"

    Returns:
        Ruta del archivo escrito.
    """
    if path_salida is None:
        path_salida = path_excel
    elif path_salida != path_excel:
        shutil.copy2(path_excel, path_salida)

    print(f"Abriendo {path_salida} para escritura...")
    wb = load_workbook(path_salida)

    nombres_hojas = wb.sheetnames
    print(f"Hojas disponibles: {nombres_hojas}")

    # --- Tramos ---
    if 'Tramos' in nombres_hojas:
        _escribir_tramos(wb['Tramos'], df_salida)
    else:
        ws = wb.create_sheet('Tramos')
        _escribir_tramos(ws, df_salida)

    # --- Resumen Tramos ---
    if 'Resumen Tramos' in nombres_hojas:
        _escribir_resumen_tramos(wb['Resumen Tramos'], df_agregado)
    else:
        ws = wb.create_sheet('Resumen Tramos')
        _escribir_resumen_tramos(ws, df_agregado)

    # --- consolidado2 ---
    if 'consolidado2' in nombres_hojas:
        _escribir_consolidado2(wb['consolidado2'], df_salida)
    else:
        ws = wb.create_sheet('consolidado2')
        _escribir_consolidado2(ws, df_salida)

    # --- Copia de Consolidado ---
    if 'Copia de Consolidado' in nombres_hojas:
        _escribir_copia_consolidado(wb['Copia de Consolidado'], df_salida)
    else:
        ws = wb.create_sheet('Copia de Consolidado')
        _escribir_copia_consolidado(ws, df_salida)

    # --- Variables del Modelo ---
    if variables is not None and 'Variables del Modelo' in nombres_hojas:
        _escribir_variables_modelo(wb['Variables del Modelo'], variables)

    wb.save(path_salida)
    print(f"\nArchivo guardado: {path_salida}")
    return path_salida


def escribir_resultados_bytes(
    xlsx_bytes: bytes,
    df_salida: pd.DataFrame,
    df_agregado: pd.DataFrame,
    suffix: str = ".xlsx",
    variables: dict = None,
) -> bytes:
    """
    Escribe Tramos, Resumen Tramos, consolidado2 y Copia de Consolidado en una
    copia en memoria del mismo libro que subió el usuario (mismo contenido base).
    Opcionalmente actualiza la hoja "Variables del Modelo" con las variables proporcionadas.

    Args:
        xlsx_bytes: bytes del archivo Excel original
        df_salida: DataFrame con resultado por tramo × año
        df_agregado: DataFrame con resumen por tramo
        suffix: sufijo del archivo
        variables: dict con variables del modelo

    Returns:
        Bytes del .xlsx completo tras guardar.
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(xlsx_bytes)
        path = tmp.name
    try:
        escribir_resultados(path, df_salida, df_agregado, path_salida=path, variables=variables)
        return Path(path).read_bytes()
    finally:
        Path(path).unlink(missing_ok=True)
